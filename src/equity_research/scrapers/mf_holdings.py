"""Monthly scheme-portfolio (holdings) scrapers — the per-AMC piece.

SEBI mandates every scheme disclose its full month-end portfolio, but there is **no
consolidated primary feed** — each AMC posts its own file (usually an XLSX). The good
news: most follow SEBI's prescribed layout (Name of the Instrument · ISIN ·
Industry/Rating · Quantity · Market/Fair Value · % to Net Assets), so a single
**header-detecting generic parser** handles the bulk; each AMC is then just a URL +
a couple of hints. Coverage grows one AMC at a time via ``REGISTRY``.

Row shape: ``{fund_name, isin, instrument, industry, quantity, market_value_cr, pct_nav}``.
"""

from __future__ import annotations

import calendar
import io
import re
from datetime import date

import openpyxl

from equity_research.common.http import fetch_bytes

_ISIN = re.compile(r"^IN[A-Z0-9]{10}$")
_MONTHS = ["", "January", "February", "March", "April", "May", "June",
           "July", "August", "September", "October", "November", "December"]


def _month_end(d: date) -> date:
    return date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])


def _load_xlsx(raw: bytes) -> openpyxl.Workbook:
    """Load an XLSX from bytes (some AMCs serve xlsx bytes under a .xls name; openpyxl's
    extension gate is bypassed by handing it a BytesIO)."""
    return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)


def _num(v) -> float | None:
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


# ---------------- generic SEBI-format parser ----------------
def _match_col(text: str) -> str | None:
    """Map a header-cell string to one of our logical columns."""
    t = text.lower()
    if "isin" in t:
        return "isin"
    if "% to net" in t or "% of net" in t or "% to nav" in t or "percentage to net" in t \
            or ("% " in t and "net asset" in t):
        return "pct"
    if "name of the instrument" in t or "name of instrument" in t or t.strip() == "instrument":
        return "name"
    if "industry" in t or "rating" in t:
        return "industry"
    if "quantity" in t:
        return "quantity"
    if "market" in t or "fair value" in t:
        return "mv"
    return None


def _header_map(row) -> dict[str, int]:
    m: dict[str, int] = {}
    for idx, c in enumerate(row or []):
        if c is None:
            continue
        col = _match_col(str(c))
        if col and col not in m:
            m[col] = idx
    return m


def _find_header(rows) -> tuple[int, dict[str, int]]:
    """First row that carries an ISIN column + a name/instrument column → (idx, colmap)."""
    for i, r in enumerate(rows[:25]):
        m = _header_map(r)
        if "isin" in m and "name" in m:
            return i, m
    return -1, {}


def _title(rows, header_idx: int, sheet: str) -> str:
    """Scheme title = first meaningful text cell above the header band."""
    for r in rows[:header_idx or 6]:
        for c in (r or [])[:4]:
            t = str(c).strip() if c else ""
            if len(t) > 6 and re.search(r"[A-Za-z]{3}", t) and "portfolio" not in t.lower() \
                    and "name of" not in t.lower():
                return t
    return sheet


def _mv_to_cr(v: float | None, header_text: str) -> float | None:
    if v is None:
        return None
    h = header_text.lower()
    if "lakh" in h:
        return v / 100
    if "crore" in h or "cr." in h or "(rs. in cr" in h:
        return v
    return v / 100          # SEBI default disclosure unit is ₹ lakh


def parse_generic(wb: openpyxl.Workbook) -> list[dict]:
    """Parse any SEBI-format monthly-portfolio workbook (one or many scheme sheets)."""
    out: list[dict] = []
    for sheet in wb.sheetnames:
        rows = list(wb[sheet].iter_rows(values_only=True))
        hi, cm = _find_header(rows)
        if hi < 0 or "pct" not in cm:
            continue
        fund_name = _title(rows, hi, sheet)
        mv_hdr = str(rows[hi][cm["mv"]]) if "mv" in cm and rows[hi][cm["mv"]] else ""
        sheet_rows: list[dict] = []
        for r in rows[hi + 1:]:
            if not r or len(r) <= cm["isin"]:
                continue
            isin = str(r[cm["isin"]]).strip() if r[cm["isin"]] else ""
            if not _ISIN.match(isin):
                continue
            sheet_rows.append({
                "fund_name": fund_name, "isin": isin,
                "instrument": (str(r[cm["name"]]).strip() if r[cm["name"]] else ""),
                "industry": (str(r[cm["industry"]]).strip() if cm.get("industry") is not None
                             and len(r) > cm["industry"] and r[cm["industry"]] else ""),
                "quantity": _num(r[cm["quantity"]]) if cm.get("quantity") is not None
                            and len(r) > cm["quantity"] else None,
                "market_value_cr": _mv_to_cr(_num(r[cm["mv"]]), mv_hdr) if cm.get("mv") is not None
                                   and len(r) > cm["mv"] else None,
                "pct_nav": _num(r[cm["pct"]]),
            })
        # per-sheet: %NAV disclosed as a fraction (sums to ~1) vs a percentage (~100)
        pcts = [x["pct_nav"] for x in sheet_rows if x["pct_nav"] is not None]
        if pcts and max(pcts) <= 1.5:
            for x in sheet_rows:
                if x["pct_nav"] is not None:
                    x["pct_nav"] *= 100
        out.extend(sheet_rows)
    return out


def _fetch_xlsx(url: str) -> tuple[str, list[dict]]:
    """Download an XLSX at ``url`` and parse it with the generic SEBI parser."""
    try:
        raw = fetch_bytes(url, timeout=60)
    except Exception:  # noqa: BLE001 — month may not be published yet
        return url, []
    try:
        return url, parse_generic(_load_xlsx(raw))
    except Exception:  # noqa: BLE001 — layout surprise; degrade, don't crash
        return url, []


def _capture_links(page_url: str, pattern: str, wait_ms: int = 8000) -> list[str]:
    """Load a JS download page in the browser tier and return anchor hrefs matching
    ``pattern`` (a JS regex source). Used for AMCs that list files client-side."""
    from scrapling.fetchers import StealthyFetcher
    js = ("() => { const re=new RegExp(" + _repr_js(pattern) + ",'i'); "
          "return JSON.stringify([...new Set([...document.querySelectorAll('a[href]')]"
          ".map(a=>a.href).filter(h=>re.test(h)))]); }")
    cap: dict = {}

    def act(page):
        try:
            page.wait_for_timeout(wait_ms)
        except Exception:  # noqa: BLE001
            pass
        cap["u"] = page.evaluate(js)
        return page

    try:
        StealthyFetcher.fetch(page_url, headless=True, network_idle=True, page_action=act)
        import json
        return json.loads(cap.get("u", "[]"))
    except Exception:  # noqa: BLE001
        return []


def _repr_js(s: str) -> str:
    """JS string literal for a regex source (escape backslashes/quotes)."""
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def _browser_pick_and_fetch(page_url: str, link_pattern: str, pick, wait_ms: int = 6000
                            ) -> tuple[str, bytes]:
    """ONE browser session: open ``page_url``, collect anchor hrefs matching ``link_pattern``
    (a JS regex source), let ``pick(links)`` choose one Python-side, then download THAT file
    *in-page* via ``fetch()`` — which carries the tab's anti-bot clearance + same-origin
    cookies, so it beats file servers that 503 a plain HTTP client. Doing both in a single
    session avoids a second page load (which some AMC servers rate-limit with a 503).
    Returns ``(chosen_url, bytes)`` — bytes ``b''`` on any failure. The file must be
    same-origin as ``page_url`` (else the in-page fetch is CORS-blocked)."""
    import base64
    import json
    from scrapling.fetchers import StealthyFetcher
    js_links = ("() => { const re=new RegExp(" + _repr_js(link_pattern) + ",'i'); "
                "return JSON.stringify([...new Set([...document.querySelectorAll('a[href]')]"
                ".map(a=>a.href).filter(h=>re.test(h)))]); }")
    js_fetch = ("async (u) => { const r=await fetch(u); if(!r.ok) return 'ERR '+r.status; "
                "const b=new Uint8Array(await r.arrayBuffer()); let s=''; "
                "for (let i=0;i<b.length;i++) s+=String.fromCharCode(b[i]); return btoa(s); }")
    cap: dict = {}

    def act(page):
        try:
            page.wait_for_timeout(wait_ms)
        except Exception:  # noqa: BLE001
            pass
        try:
            links = json.loads(page.evaluate(js_links))
        except Exception:  # noqa: BLE001
            links = []
        url = pick(links) if links else None
        cap["url"] = url or ""
        cap["d"] = page.evaluate(js_fetch, url) if url else ""
        return page

    try:
        StealthyFetcher.fetch(page_url, headless=True, network_idle=True, page_action=act)
    except Exception:  # noqa: BLE001
        return "", b""
    url, d = cap.get("url") or "", cap.get("d") or ""
    if not d or d.startswith("ERR"):
        return url, b""
    try:
        return url, base64.b64decode(d)
    except Exception:  # noqa: BLE001
        return url, b""


# ---------------- per-AMC URL builders ----------------
_PPFAS_URL = ("https://amc.ppfas.com/downloads/portfolio-disclosure/"
              "{y}/PPFAS_Monthly_Portfolio_Report_{mon}_{d}_{y}.xls")


def fetch_ppfas(as_of: date) -> tuple[str, list[dict]]:
    """PPFAS month-end holdings across all its schemes (one workbook, sheet per scheme)."""
    me = _month_end(as_of)
    return _fetch_xlsx(_PPFAS_URL.format(y=me.year, mon=_MONTHS[me.month], d=me.day))


_HDFC_PAGE = "https://www.hdfcfund.com/statutory-disclosure/monthly-portfolio"
_HDFC_FILE = re.compile(r"files\.hdfcfund\.com/.*Monthly.*\.xlsx?$", re.I)


def fetch_hdfc(as_of: date) -> tuple[str, list[dict]]:
    """HDFC posts one file per scheme on a direct CDN; the monthly page lists them all.
    Capture the file links, then parse each with the generic parser."""
    me = _month_end(as_of)
    tag = f"{me.day:02d} {_MONTHS[me.month][:3]}"        # e.g. '31 May' — keep only this month's files
    links = [ln for ln in _capture_links(_HDFC_PAGE, r"files\.hdfcfund\.com/.*Monthly.*\.xlsx?")
             if _HDFC_FILE.search(ln)]
    from urllib.parse import unquote
    links = [ln for ln in links if tag.lower() in unquote(ln).lower()]
    rows: list[dict] = []
    for ln in links:
        rows.extend(_fetch_xlsx(ln)[1])
    return _HDFC_PAGE, rows


# Nippon posts ONE consolidated month-end workbook (sheet per scheme) on its disclosures
# page, but the file server 503s a plain client — so we capture the link and fetch it
# from inside the browser. Filenames vary in month spelling (Jun vs June), so match by date.
_NIPPON_PAGE = ("https://mf.nipponindiaim.com/investor-service/downloads/"
                "factsheet-portfolio-and-other-disclosures")


def _nippon_link_date(ln: str) -> date | None:
    m = re.search(r"MONTHLY-PORTFOLIO-(\d{1,2})-([A-Za-z]+)-(\d{2})\.xlsx?", ln, re.I)
    if not m:
        return None
    mon = {mn[:3].lower(): i for i, mn in enumerate(_MONTHS) if mn}.get(m.group(2)[:3].lower())
    if not mon:
        return None
    try:
        return date(2000 + int(m.group(3)), mon, int(m.group(1)))
    except ValueError:
        return None


def fetch_nippon(as_of: date) -> tuple[str, list[dict]]:
    """Nippon India month-end holdings (one consolidated workbook, sheet per scheme)."""
    me = _month_end(as_of)

    def pick(links: list[str]) -> str | None:
        dated = [(_nippon_link_date(ln), ln) for ln in links if _nippon_link_date(ln)]
        exact = [ln for d, ln in dated if d == me]
        return exact[0] if exact else (max(dated)[1] if dated else None)

    url, raw = _browser_pick_and_fetch(_NIPPON_PAGE, r"NIMF-MONTHLY-PORTFOLIO", pick, wait_ms=7000)
    return (url, parse_generic(_load_xlsx(raw))) if raw else (url, [])


# AMC display-name (as in ``mf_scheme.amc``) -> fetcher(as_of) -> (url, rows)
REGISTRY = {
    "PPFAS Mutual Fund": fetch_ppfas,
    "HDFC Mutual Fund": fetch_hdfc,
    "Nippon India Mutual Fund": fetch_nippon,
}


def fetch_amc_holdings(amc_name: str, as_of: date) -> tuple[str, list[dict]]:
    """Holdings for a registered AMC at ``as_of`` (empty if the AMC isn't covered)."""
    fn = REGISTRY.get(amc_name)
    return fn(as_of) if fn else ("", [])
